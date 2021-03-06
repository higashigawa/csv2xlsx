# モジュールのインポート
import os, tkinter, tkinter.filedialog, tkinter.messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import subprocess


# ファイル選択ダイアログの表示
root = tkinter.Tk()
root.withdraw()
fTyp = [("全てのファイル","*.*"), ("csvファイル","*.csv"), ("datファイル","*.dat"), ("txtファイル","*.txt")]
iDir = os.path.abspath(os.path.dirname(__file__))
tkinter.messagebox.showinfo('エクセル変換プログラム','処理ファイルを選択してください！')

file1 = tkinter.filedialog.askopenfilename(filetypes = fTyp,initialdir = iDir)
file2 = file1[:-3]+'xlsx'

# 変換処理
try:
    df = pd.read_csv(file1, encoding='cp932', header=None, dtype="object")

except:
    df = pd.read_csv(file1, encoding='cp932', header=0, dtype="object")
    file1a = file1+'_data.txt'
    df.to_csv(file1a, encoding='cp932', header=False)
    df = pd.read_csv(file1a, encoding='cp932', header=None, dtype="object")

    # 一時ファイル削除
    os.remove(file1a)
    
    # エクセルファイル書き出し
    df.to_excel(file2, encoding='cp932', index=None, header=None)

    # 先頭行挿入    
    wb = openpyxl.load_workbook(file2)
    #ws = wb.active
    ws = wb.worksheets[0]
    ws.insert_rows(0)
    ws['A1'].value = len(df)
    wb.save(file2)

else:
    # エクセルファイル書き出し
    df.to_excel(file2, encoding='cp932', index=None, header=None)

# read input xlsx
wb1 = openpyxl.load_workbook(file2)
ws1 = wb1.worksheets[0]

# set font
font = Font(name='メイリオ', size=10)

# write in sheet
for row in ws1:
    for cell in row:
        ws1[cell.coordinate].font = font

# save xlsx file
wb1.save(file2)

# 処理ファイル名の出力
files = '変換前 '+file1+'\n変換後 '+file2
tkinter.messagebox.showinfo('エクセル変換プログラム',files)

# Excelを起動する
try:
    subprocess.Popen([r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE', file2])
except:
    subprocess.Popen([r'C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE', file2])
else:
    pass
